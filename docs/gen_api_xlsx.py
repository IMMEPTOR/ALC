from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()
ws = wb.active
ws.title = "API"

headers = ["Роут", "Заголовок", "Цель (описание)", "Что принимает", "Что возвращает", "Доп. статус-коды"]

# Styles
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_align = Alignment(vertical="top", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# Write header
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Data
rows = [
    # Auth
    [
        "POST /api/auth/login",
        "Вход в систему",
        "Аутентифицирует пользователя по логину и паролю, создаёт JWT-сессию и возвращает токен",
        "Body: { username: string, password: string }",
        "{ token: string, user: { id, username, role, permissions[] } }",
        "400 — не указаны username/password\n401 — неверные учётные данные\n403 — аккаунт деактивирован",
    ],
    [
        "GET /api/auth/me",
        "Текущий пользователь",
        "Возвращает данные авторизованного пользователя по токену",
        "Header: Authorization: Bearer <token>",
        "{ user: { id, username, role } }",
        "401 — не авторизован",
    ],
    [
        "POST /api/auth/logout",
        "Выход из системы",
        "Удаляет серверную сессию, инвалидирует токен",
        "Header: Authorization: Bearer <token>",
        '{ message: "Выход выполнен" }',
        "401 — не авторизован",
    ],

    # Users
    [
        "GET /api/users",
        "Список пользователей",
        "Получает список всех пользователей с их ролями (без пароля). Только admin",
        "Header: Authorization: Bearer <token>",
        "User[] (массив объектов с populated role_id)",
        "401 — не авторизован\n403 — нет прав (не admin)",
    ],
    [
        "GET /api/users/:id",
        "Пользователь по ID",
        "Получает одного пользователя по идентификатору. Только admin",
        "Path: id — идентификатор пользователя",
        "User объект (populated role_id, без password_hash)",
        "401\n403\n404 — пользователь не найден",
    ],
    [
        "POST /api/users",
        "Создание пользователя",
        "Создаёт нового пользователя с указанной ролью. Только admin",
        "Body: { username: string, password: string, role_name: string }",
        "{ id: string, username: string }",
        "400 — обязательные поля не заполнены\n409 — пользователь уже существует\n401, 403",
    ],
    [
        "PUT /api/users/:id",
        "Обновление пользователя",
        "Обновляет данные пользователя (роль, активность, пароль). Только admin",
        "Path: id\nBody: { is_active?: boolean, role_name?: string, password?: string }",
        "User обновлённый объект",
        "400 — роль не найдена\n404 — пользователь не найден\n401, 403",
    ],
    [
        "DELETE /api/users/:id",
        "Удаление пользователя",
        "Удаляет пользователя из системы. Только admin",
        "Path: id — идентификатор пользователя",
        '{ message: "Пользователь удален" }',
        "404 — пользователь не найден\n401, 403",
    ],

    # Sites
    [
        "GET /api/sites",
        "Список площадок",
        "Получает все производственные площадки",
        "Header: Authorization: Bearer <token>",
        "ProductionSite[]",
        "401",
    ],
    [
        "GET /api/sites/:id",
        "Площадка по ID",
        "Получает одну производственную площадку по идентификатору",
        "Path: id",
        "ProductionSite объект",
        "401\n404 — площадка не найдена",
    ],
    [
        "POST /api/sites",
        "Создание площадки",
        "Создаёт новую производственную площадку",
        "Body: { name: string, location: string }",
        "ProductionSite объект",
        "400 — обязательные поля не заполнены\n401",
    ],
    [
        "PUT /api/sites/:id",
        "Обновление площадки",
        "Обновляет данные производственной площадки",
        "Path: id\nBody: { name?: string, location?: string }",
        "ProductionSite обновлённый объект",
        "404 — площадка не найдена\n401",
    ],
    [
        "DELETE /api/sites/:id",
        "Удаление площадки",
        "Удаляет производственную площадку",
        "Path: id",
        '{ message: "Площадка удалена" }',
        "404\n401",
    ],

    # Lines
    [
        "GET /api/lines",
        "Список линий",
        "Получает все сборочные линии с возможностью фильтрации по площадке",
        "Query: ?site_id=<id> (опционально)",
        "AssemblyLine[] (populated site_id)",
        "401",
    ],
    [
        "GET /api/lines/:id",
        "Линия по ID",
        "Получает одну сборочную линию по идентификатору",
        "Path: id",
        "AssemblyLine (populated site_id)",
        "401\n404 — линия не найдена",
    ],
    [
        "POST /api/lines",
        "Создание линии",
        "Создаёт новую сборочную линию для указанной площадки",
        "Body: { site_id: string, name: string, status?: string }",
        "AssemblyLine объект",
        "400 — обязательные поля не заполнены\n401",
    ],
    [
        "PUT /api/lines/:id",
        "Обновление линии",
        "Обновляет данные сборочной линии (имя, статус)",
        "Path: id\nBody: { name?: string, status?: string }",
        "AssemblyLine обновлённый объект",
        "404\n401",
    ],
    [
        "DELETE /api/lines/:id",
        "Удаление линии",
        "Удаляет сборочную линию",
        "Path: id",
        '{ message: "Линия удалена" }',
        "404\n401",
    ],

    # Nodes
    [
        "GET /api/nodes",
        "Список узлов",
        "Получает все технологические узлы с фильтрацией по линии и статусу",
        "Query: ?line_id=<id>&status=<online|offline|warning|critical> (опционально)",
        "TechNode[] (populated line_id)",
        "401",
    ],
    [
        "GET /api/nodes/:id",
        "Узел по ID",
        "Получает один технологический узел с его параметрами",
        "Path: id",
        "TechNode (populated line_id, включает parameters[])",
        "401\n404 — узел не найден",
    ],
    [
        "POST /api/nodes",
        "Создание узла",
        "Создаёт новый технологический узел с набором контролируемых параметров",
        "Body: { line_id: string, name: string, type: string, ip_address: string, parameters?: Parameter[] }",
        "TechNode объект",
        "400 — обязательные поля не заполнены\n401",
    ],
    [
        "PUT /api/nodes/:id",
        "Обновление узла",
        "Обновляет данные узла (статус, параметры и т.д.)",
        "Path: id\nBody: { name?, type?, status?, ip_address?, parameters? }",
        "TechNode обновлённый объект",
        "404\n401",
    ],
    [
        "DELETE /api/nodes/:id",
        "Удаление узла",
        "Удаляет технологический узел",
        "Path: id",
        '{ message: "Узел удален" }',
        "404\n401",
    ],

    # Telemetry
    [
        "GET /api/telemetry",
        "История телеметрии",
        "Получает записи телеметрии с фильтрацией по узлу, параметру и временному диапазону",
        "Query: ?node_id=<id>&param_id=<id>&from=<ISO>&to=<ISO>&limit=<N> (все опционально, limit по умолч. 100)",
        "TelemetryRecord[] (отсортировано по timestamp desc)",
        "401",
    ],
    [
        "GET /api/telemetry/latest/:node_id",
        "Последняя телеметрия узла",
        "Получает последнее значение каждого параметра указанного узла",
        "Path: node_id — идентификатор узла",
        "LatestTelemetry[] — массив { param_id, name, unit, min_value, max_value, value, timestamp, quality_flag }",
        "401\n404 — узел не найден",
    ],
    [
        "POST /api/telemetry",
        "Запись телеметрии",
        "Записывает новое значение телеметрии. Автоматически проверяет пороги параметра и при превышении создаёт Alert, обновляет статус узла",
        "Body: { node_id: string, param_id: string, value: number, quality_flag?: string }",
        "TelemetryRecord объект",
        "400 — обязательные поля не заполнены\n401",
    ],

    # Alerts
    [
        "GET /api/alerts",
        "Список алармов",
        "Получает алармы с фильтрацией по узлу, статусу и уровню критичности",
        "Query: ?node_id=<id>&status=<active|acknowledged|resolved>&severity=<info|warning|critical>&limit=<N> (по умолч. 50)",
        "Alert[] (populated node_id, отсортировано по created_at desc)",
        "401",
    ],
    [
        "GET /api/alerts/:id",
        "Аларм по ID",
        "Получает один аларм по идентификатору",
        "Path: id",
        "Alert (populated node_id)",
        "401\n404 — аларм не найден",
    ],
    [
        "PATCH /api/alerts/:id/acknowledge",
        "Подтверждение аларма",
        "Подтверждает аларм — переводит статус в acknowledged. Доступно: operator, engineer, admin",
        "Path: id",
        "Alert обновлённый (status: acknowledged)",
        "401\n403 — нет прав\n404 — аларм не найден",
    ],
    [
        "PATCH /api/alerts/:id/resolve",
        "Разрешение аларма",
        "Разрешает аларм — переводит статус в resolved, заполняет resolved_at. Доступно: engineer, admin",
        "Path: id",
        "Alert обновлённый (status: resolved, resolved_at: Date)",
        "401\n403 — нет прав\n404 — аларм не найден",
    ],

    # Commands
    [
        "GET /api/commands",
        "Список команд",
        "Получает историю команд управления с фильтрацией по узлу и статусу",
        "Query: ?node_id=<id>&status=<pending|executing|completed|failed>&limit=<N> (по умолч. 50)",
        "Command[] (populated node_id, user_id, отсортировано по created_at desc)",
        "401",
    ],
    [
        "GET /api/commands/:id",
        "Команда по ID",
        "Получает одну команду управления по идентификатору",
        "Path: id",
        "Command (populated node_id, user_id)",
        "401\n404 — команда не найдена",
    ],
    [
        "POST /api/commands",
        "Отправка команды",
        "Создаёт команду управления для узла. Автоматически выполняется через ~1 сек. Доступно: operator, engineer, admin",
        "Body: { node_id: string, action_type: string, parameters?: object }",
        "Command объект (status: pending)",
        "400 — обязательные поля не заполнены\n401\n403 — нет прав",
    ],

    # Health
    [
        "GET /api/health",
        "Проверка здоровья",
        "Проверяет доступность сервера",
        "—",
        '{ status: "ok", timestamp: string }',
        "—",
    ],
]

# Write data
for r, row_data in enumerate(rows, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.alignment = cell_align
        cell.border = thin_border

# Column widths
ws.column_dimensions["A"].width = 38
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 50
ws.column_dimensions["D"].width = 48
ws.column_dimensions["E"].width = 48
ws.column_dimensions["F"].width = 35

# Freeze header
ws.freeze_panes = "A2"

wb.save("/Users/yroslav/Desktop/alc/docs/API.xlsx")
print("OK")
